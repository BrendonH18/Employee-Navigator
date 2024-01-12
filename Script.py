import csv
from openpyxl import Workbook
import pandas as pd
import os, re, io
from pathlib import Path
import PySimpleGUI as ui
from multiprocessing import Process
import functools

def choose_disability() -> dict:
    select_file = [
        [ui.Checkbox("Long Term", key='Long Term Disability')],
        [ui.Checkbox("Short Term", key='Short Term Disability')],
        [ui.Button("Start")],
    ]
    window = ui.Window('SteveConnections', select_file)
    file_exists = False
    while not file_exists:
        event, values = window.read()
        if event == ui.WINDOW_CLOSED:
            window.close()
            return values
            plan_types = []
            for key, value in values.items():
                if value == True:
                    plan_types.append(key)
            return plan_types
        elif event == "Start":
            window.close()
            return values
            plan_types = []
            for key, value in values.items():
                if value == True:
                    plan_types.append(key)
            window.close()
            return plan_types

def choose_file(text: str = "File: ") -> str:
    select_file = [
        [ui.Text(text), ui.InputText(key='-file1-'), ui.FileBrowse()],
        [ui.Button("Start")],
    ]
    window = ui.Window('SteveConnections', select_file)
    file_exists = False
    while not file_exists:
        event, values = window.read()
        if event == ui.WINDOW_CLOSED:
            break
        elif event == "Start":
            filename = values['-file1-']
            while True:
                if not Path(filename).is_file():
                    if filename == '':
                        ui.popup_ok('Please select a file!')
                    else:
                        ui.popup_ok("That file doesn't exist!")
                    filename = ui.popup_get_file("", no_window=True)
                    if filename == '':
                        break
                    window['-file1-'].update(filename)
                else:
                    print('File is ready !')
                    file_exists = True
                    break
    window.close()
    return str(values['-file1-'])

class SocialSecurity():
    def __init__(self, reference_dict=None):
        self.reference_dict = reference_dict if reference_dict is not None else {}
        self.ssn_generator = self.generate_ssn()

    @staticmethod
    def generate_ssn(start=111111111):
        while True:
            ssn = f"{start:09d}"
            yield f"{ssn[:3]}-{ssn[3:5]}-{ssn[5:]}"
            start += 1

    def storeValues(self, df: pd.DataFrame):
        for index, row in df.iterrows():
            reference = (row['EID'], row['First Name'], row['Last Name'], row['Relationship'])
            if reference not in self.reference_dict and row['SSN']:
                self.reference_dict[reference] = row['SSN']

    def updateSsn(self, df: pd.DataFrame):
        for index, row in df.iterrows():
            if pd.isna(row['SSN']):
                reference = (row['EID'], row['First Name'], row['Last Name'], row['Relationship'])
                if reference in self.reference_dict:
                    df.at[index, 'SSN'] = self.reference_dict[reference]
                else:
                    new_ssn = next(self.ssn_generator)
                    while True:
                        if df['SSN'].str.contains(new_ssn).any():
                            new_ssn = next(self.ssn_generator)
                            continue
                        break
                    df.at[index, 'SSN'] = new_ssn
                    self.reference_dict[reference] = new_ssn
        return df
        
ssn = SocialSecurity()
def csv_to_excel(csv_file, filename:str) -> Workbook:
    # csv_data = []
    # with open(csv_file) as file_obj:
    #     reader = csv.reader(file_obj)
    #     for row in reader:
    #         csv_data.append(row)
    # wb = Workbook()
    # sheet = wb.active
    # wb.headers = csv_data[0]
    # for row in csv_data:
    #     sheet.append(row)
    # with pd.ExcelFile(wb) as xl:
    #     sheetnames = xl.sheet_names
    #     df = xl.parse(sheetnames[0])
    #     ssn.updateSsn(df=df)
    # output = io.BytesIO()
    # return_value = df.to_excel(output, index=False)
    # return return_value
    df = pd.read_csv(csv_file)
    ssn.updateSsn(df=df)
    remove_file(f"./{filename}")
    df.to_excel(f"{filename}", index=False)

class SpecialFunctions():
    def __init__(self, previous_element = None, previous_employee_row_number = 0, row_number = 0) -> None:
        self.previous_element = previous_element
        self.previous_employee_row_number = previous_employee_row_number
        self.row_number = row_number

    def return_X_if_Column_Equals(self, row, return_value: str, column: str, equals_value: str, isRequestReset: bool = False):
        # x = previous_element
        if isRequestReset:
            self.previous_element = None
        if row[column] == equals_value:
            self.previous_element = row[return_value]
            return self.previous_element
        return self.previous_element
    def add_value_to_previous_employee(self, row: pd.DataFrame, relationship = "Spouse"):
        if row["Relationship"].iloc[0] == "Employee":
            self.previous_employee_row_number = self.row_number
        self.row_number += 1
        return self.previous_employee_row_number
    

        

def remove_file(file: str) -> None:
    if os.path.exists(file):
            os.remove(file)




def main_basic_benefits_census(wb: str, file_name_suffix: str = "Base"):
    col_titles = ["EID", "SSN", "Relationship", "Last Name","First Name", "Plan Display Name", "Effective Date", "Coverage Tier", "Job Class"]
    with pd.ExcelFile(wb) as xl:
        sheetnames = xl.sheet_names
        df = xl.parse(sheetnames[0])
        filtered_df = df[(df["Coverage Tier"] != "Waived") & 
                        (
                            (df["Plan Type"] == "Medical") |
                            (df["Plan Type"] == "Dental") |
                            (df["Plan Type"] == "Vision") 
                        )][col_titles]
        sf = SpecialFunctions()
        filtered_df["Employee SSN"] = filtered_df.apply(functools.partial(sf.return_X_if_Column_Equals, return_value="SSN", column="Relationship", equals_value="Employee"), axis=1)
        filtered_df["Job Class"] = filtered_df.apply(functools.partial(sf.return_X_if_Column_Equals, return_value="Job Class", column="Relationship", equals_value="Employee"), axis=1)
        filtered_df.rename(columns={"EID": "Employee ID","Job Class" : "Class", "SSN": "Dependent SSN", "Effective Date": "EE Effective Plan Start Date", "Plan Display Name": "Plan Name"}, inplace=True)
        filter_new_order = ["Employee ID", "Employee SSN", "Dependent SSN", "Relationship", "Last Name", "First Name", "Plan Name", "EE Effective Plan Start Date", "Class"]
        filtered_df = filtered_df[filter_new_order]
        filtered_df['Class'] = filtered_df['Class'].fillna('')
        filtered_df['Class'] = filtered_df['Class'].replace('', 'Default')
        filtered_df = filtered_df.dropna(axis=1, how="all")
        remove_file(f"./{file_name_suffix}.xlsx")
        filtered_df.to_excel(f"{file_name_suffix}.xlsx", index=False, sheet_name=f"{file_name_suffix}")

def main_employee_dependent_census(wb: str, file_name_suffix: str = "Base"):
    col_titles = [
        "Location", "EID",
        "First Name", "Middle Name", "Last Name",
        "Relationship",
        "SSN", "Sex", "Birth Date", "Address 1", "Address 2", "City", "State", "Zip",
        "Personal Phone", "Work Phone", "Email", "Personal Email",
        "Employee Type", "Employee Status", "Hire Date", "Termination Date", "Termination Type",
        "Job Class", "Job Title", "Compensation", "Compensation Type", "Scheduled Hours"]
    with pd.ExcelFile(wb) as xl:
        sheetnames = xl.sheet_names
        df = xl.parse(sheetnames[0])

        filtered_df = df[col_titles]
        #Add Employee SSN
        sf = SpecialFunctions()
        sf1 = SpecialFunctions()
        sf2 = SpecialFunctions()
        filtered_df["Employee SSN"] = filtered_df.apply(functools.partial(sf.return_X_if_Column_Equals, return_value="SSN", column="Relationship", equals_value="Employee"), axis=1)
        filtered_df["Job Class"] = filtered_df.apply(functools.partial(sf1.return_X_if_Column_Equals, return_value="Job Class", column="Relationship", equals_value="Employee"), axis=1)
        filtered_df["Is Full Time"] = filtered_df.apply(functools.partial(sf2.return_X_if_Column_Equals, return_value="Employee Type", column="Relationship", equals_value="Employee"), axis=1)
        filtered_df["Is Full Time"] = filtered_df["Is Full Time"].fillna("No").replace(to_replace={"Full-Time": "Yes", "Part-Time": "No", "Contractor": "No"})
        #add column
        filtered_df["Salary Effective Date"] = filtered_df["Hire Date"]
        # print(filtered_df["Compensation Type"].name)
        filtered_df["Annual Base Salary"] = filtered_df.apply(lambda x: float(x["Compensation"].replace("$","").replace(",","")) if x["Compensation Type"] == "Salary" else None, axis=1)
        filtered_df["Hourly Rate"] = filtered_df.apply(lambda x: float(x["Compensation"].replace("$","").replace(",","")) if x["Compensation Type"] == "Hourly" else None, axis=1)
        filtered_df["Hours Per Week"] = filtered_df.apply(lambda x: float(x["Scheduled Hours"]) if x["Compensation Type"] == "Hourly" else None, axis=1)

        #renames
        filtered_df.rename(columns={"EID": "Employee ID", "SSN": "Dependent SSN", "Birth Date": "Date of Birth", "Job Class": "Class", "Location" : "Office", "Email" :"Work Email", "Zip": "Zip Code", "Personal Phone" : "Phone Number", "Termination Type": "Termination Reason"}, inplace=True)
    
        filter_new_order = ["Employee ID","Employee SSN", "Dependent SSN", "Relationship", "Last Name", "First Name", "Middle Name", "Date of Birth", "Sex", "Hire Date", "Class", "Office", "Work Email", "Personal Email", "Annual Base Salary", "Hourly Rate", "Hours Per Week", "Salary Effective Date", "Address 1", "Address 2", "City", "State", "Zip Code", "Phone Number", "Job Title", "Is Full Time", "Termination Date", "Termination Reason"]
        # filter_new_order = ["Annual Base Salary", "Hourly Rate", "Hours Per Week", "Employee SSN", "Dependent SSN", "Relationship", "Last Name", "First Name", "Middle Name", "Date of Birth", "Sex", "Hire Date", "Class", "Office", "Work Email", "Personal Email", "Salary Effective Date", "Address 1", "Address 2", "City", "State", "Zip Code", "Phone Number", "Job Title", "Is Full Time", "Termination Date", "Termination Reason"]
        filtered_df = filtered_df[filter_new_order]
        filtered_df['Class'] = filtered_df['Class'].fillna('')
        filtered_df['Class'] = filtered_df['Class'].replace('', 'Default')
        filtered_df = filtered_df.dropna(axis=1, how="all")
        remove_file(f"./{file_name_suffix}.xlsx")
        filtered_df.to_excel(f"{file_name_suffix}.xlsx", index=False, sheet_name=f"{file_name_suffix}")


def main_hsa_enrollments(wb: str, file_name_suffix: str = "Base"):
    col_titles = [
        "EID","SSN", "Relationship", "Last Name","First Name", "Plan Display Name", "Effective Date", "Election Status", "Plan Type", "Coverage Tier", "Job Class"
        ]
    with pd.ExcelFile(wb) as xl:
        sheetnames = xl.sheet_names
        df = xl.parse(sheetnames[0])

        filtered_df = df[col_titles]
        filtered_df = df[(df["Plan Type"] == "Health Savings Account") & 
                        (
                            (df["Coverage Tier"] != "Waived")
                        )]
       
        #renames
        sf = SpecialFunctions()
        filtered_df["Job Class"] = filtered_df.apply(functools.partial(sf.return_X_if_Column_Equals, return_value="Job Class", column="Relationship", equals_value="Employee"), axis=1)
        filtered_df = filtered_df.rename(columns={"EID": "Employee ID","Job Class" : "Class", "SSN": "Employee SSN", "Plan Display Name": "Plan Name", "Effective Date": "EE Effective Plan Start Date", "Election Status" : "EE Annual Contribution"})
        filtered_df["EE Annual Contribution"] = filtered_df.apply(lambda x: float(x["EE Annual Contribution"].replace("$","").replace(",","")),axis=1)

        filter_new_order = ["Employee ID","Employee SSN", "Last Name", "First Name", "Plan Name", "EE Effective Plan Start Date", "EE Annual Contribution", "Class"]
        # filter_new_order = ["Annual Base Salary", "Hourly Rate", "Hours Per Week", "Employee SSN", "Dependent SSN", "Relationship", "Last Name", "First Name", "Middle Name", "Date of Birth", "Sex", "Hire Date", "Class", "Office", "Work Email", "Personal Email", "Salary Effective Date", "Address 1", "Address 2", "City", "State", "Zip Code", "Phone Number", "Job Title", "Is Full Time", "Termination Date", "Termination Reason"]
        filtered_df = filtered_df[filter_new_order]
        filtered_df['Class'] = filtered_df['Class'].fillna('')
        filtered_df['Class'] = filtered_df['Class'].replace('', 'Default')
        filtered_df = filtered_df.dropna(axis=1, how="all")
        remove_file(f"./{file_name_suffix}.xlsx")
        filtered_df.to_excel(f"{file_name_suffix}.xlsx", index=False, sheet_name=f"{file_name_suffix}")

def main_disability(wb: str, file_name_suffix: str = "Base", plan_types = {}):
    

    group_types = []
    vol_types = []
    for key, value in plan_types.items():
        if value == True:
            group_types.append(key)
            continue
        vol_types.append(key)


    col_titles = [
        "EID","SSN", "Relationship", "Last Name","First Name", "Plan Display Name", "Effective Date", "Election Status", "Plan Type", "Coverage Tier", "Job Class"
        ]
    with pd.ExcelFile(wb) as xl:
        sheetnames = xl.sheet_names
        df = xl.parse(sheetnames[0])

        filtered_df = df[col_titles]
        filtered_df_group = filtered_df.loc[filtered_df["Plan Type"].isin(group_types)]
        filtered_df_vol = filtered_df.loc[filtered_df["Plan Type"].isin(vol_types)]


        remove_file(f"./{file_name_suffix} - Group.xlsx")
        if len(group_types) > 0:
            #renames
            sf = SpecialFunctions()
            filtered_df["Job Class"] = filtered_df.apply(functools.partial(sf.return_X_if_Column_Equals, return_value="Job Class", column="Relationship", equals_value="Employee"), axis=1)
            filtered_df_group = filtered_df_group.rename(columns={"EID": "Employee ID","Job Class":"Class", "SSN": "Employee SSN", "Plan Display Name": "Plan Name", "Effective Date": "EE Effective Plan Start Date", "Election Status" : "Plan Benefit Amount"})
            filtered_df_group["Plan Benefit Amount"] = filtered_df_group.apply(lambda x: float(x["Plan Benefit Amount"].replace("$","").replace(",","")),axis=1)
            filter_new_order = ["Employee ID","Employee SSN", "Last Name", "First Name", "Plan Name", "EE Effective Plan Start Date", "Plan Benefit Amount", "Class"]
            # filter_new_order = ["Annual Base Salary", "Hourly Rate", "Hours Per Week", "Employee SSN", "Dependent SSN", "Relationship", "Last Name", "First Name", "Middle Name", "Date of Birth", "Sex", "Hire Date", "Class", "Office", "Work Email", "Personal Email", "Salary Effective Date", "Address 1", "Address 2", "City", "State", "Zip Code", "Phone Number", "Job Title", "Is Full Time", "Termination Date", "Termination Reason"]
            filtered_df_group = filtered_df_group[filter_new_order]
            filtered_df['Class'] = filtered_df['Class'].fillna('')
            filtered_df['Class'] = filtered_df['Class'].replace('', 'Default')
            filtered_df_group = filtered_df_group.dropna(axis=1, how="all")
            filtered_df_group.to_excel(f"{file_name_suffix} - Group.xlsx", index=False, sheet_name=f"{file_name_suffix}")

        remove_file(f"./{file_name_suffix} - Voluntary.xlsx")
        if len(vol_types) > 0:
            #renames
            sf = SpecialFunctions()
            filtered_df["Job Class"] = filtered_df.apply(functools.partial(sf.return_X_if_Column_Equals, return_value="Job Class", column="Relationship", equals_value="Employee"), axis=1)
            filtered_df_vol = filtered_df_vol.rename(columns={"EID": "Employee ID","Job Class": "Class", "SSN": "Employee SSN", "Plan Display Name": "Plan Name", "Effective Date": "EE Effective Plan Start Date", "Election Status" : "Plan Benefit Amount"})
            filtered_df_vol["Plan Benefit Amount"] = filtered_df_vol.apply(lambda x: float(x["Plan Benefit Amount"].replace("$","").replace(",","")),axis=1)

            filter_new_order = ["Employee ID","Employee SSN", "Last Name", "First Name", "Plan Name", "EE Effective Plan Start Date", "Plan Benefit Amount", "Class"]
            # filter_new_order = ["Annual Base Salary", "Hourly Rate", "Hours Per Week", "Employee SSN", "Dependent SSN", "Relationship", "Last Name", "First Name", "Middle Name", "Date of Birth", "Sex", "Hire Date", "Class", "Office", "Work Email", "Personal Email", "Salary Effective Date", "Address 1", "Address 2", "City", "State", "Zip Code", "Phone Number", "Job Title", "Is Full Time", "Termination Date", "Termination Reason"]
            filtered_df_vol = filtered_df_vol[filter_new_order]
            filtered_df['Class'] = filtered_df['Class'].fillna('')
            filtered_df['Class'] = filtered_df['Class'].replace('', 'Default')
            filtered_df_vol = filtered_df_vol.dropna(axis=1, how="all")
            filtered_df_vol.to_excel(f"{file_name_suffix} - Voluntary.xlsx", index=False, sheet_name=f"{file_name_suffix}")

def main_life_add_critical_illness(wb: str, file_name_suffix: str = "Base"):
    # plan_types = choose_disability()

    col_titles = [
        "EID","SSN", "Relationship", "Last Name","First Name", "Plan Display Name", "Effective Date", "Election Status", "Plan Type", "Coverage Tier", "Carrier", "Job Class"
        ]
    with pd.ExcelFile(wb) as xl:
        sheetnames = xl.sheet_names
        df = xl.parse(sheetnames[0])

        filtered_df = df[col_titles]
        filtered_df = filtered_df.loc[~filtered_df["Carrier"].isin(["Colonial Life"])]
        filtered_df = filtered_df.loc[~filtered_df["Coverage Tier"].isin(["Waived"])]
        filtered_df = filtered_df.loc[filtered_df["Plan Type"].isin(["Life/AD&D", "Life", "AD&D", "Supplemental Life", "Voluntary Life/AD&D", "Voluntary Life", "Voluntary AD&D", "Critical Illness", "Voluntary Critical Illness", "Voluntary Term Life", "Universal Life", "Whole Life", "Term Life"])]
        
        def extract_number(s):
            # Try to find a number inside parentheses
            match = re.search(r'\(\$([\d,]+)', s)
            if match:
                return float(match.group(1).replace(",",""))
            match = re.search(r'\$([\d,]+)', s)
            if match:
                return float(match.group(1).replace(",",""))
            return 0
        filtered_df["Election Status"] = filtered_df["Election Status"].apply(extract_number)

        # filtered_df["Election Status"] = filtered_df.apply(lambda x: x["Election Status"].replace("$","").replace(",",""),axis=1)

        sf = SpecialFunctions()
        new_column = []
        for row in range(0,filtered_df.shape[0]):
            employee_row = sf.add_value_to_previous_employee(filtered_df.iloc[[row]])
            if filtered_df["Relationship"].iloc[[row]].values[0] == "Spouse":
                new_column.append(None)
                new_column[employee_row] = filtered_df["Election Status"].iloc[[row]].values[0]
                continue
            
            new_column.append(None)
        filtered_df["Spouse Benefit Amount"] = new_column

        sf1 = SpecialFunctions()
        new_column = []
        for row in range(0,filtered_df.shape[0]):
            employee_row = sf1.add_value_to_previous_employee(filtered_df.iloc[[row]])
            if filtered_df["Relationship"].iloc[[row]].values[0] == "Child":
                new_column.append(None)
                new_column[employee_row] = filtered_df["Election Status"].iloc[[row]].values[0]
                continue
            
            new_column.append(None)
        filtered_df["Child Benefit Amount"] = new_column

        sf = SpecialFunctions()
        filtered_df["Job Class"] = filtered_df.apply(functools.partial(sf.return_X_if_Column_Equals, return_value="Job Class", column="Relationship", equals_value="Employee"), axis=1)


        filtered_df = filtered_df.loc[filtered_df["Relationship"].isin(["Employee"])]
        filtered_df = filtered_df.rename(columns={"EID": "Employee ID","Job Class": "Class", "SSN": "Employee SSN", "Plan Display Name": "Plan Name", "Effective Date": "EE Effective Plan Start Date", "Election Status" : "Benefit Amount"})
        filter_new_order = ["Employee ID","Employee SSN", "Last Name", "First Name", "Plan Name", "EE Effective Plan Start Date", "Benefit Amount", "Spouse Benefit Amount",	"Child Benefit Amount", "Class"]
        filtered_df = filtered_df[filter_new_order]
        filtered_df['Class'] = filtered_df['Class'].fillna('')
        filtered_df['Class'] = filtered_df['Class'].replace('', 'Default')
        
        remove_file(f"./{file_name_suffix}.xlsx")
        filtered_df.to_excel(f"{file_name_suffix}.xlsx", index=False, sheet_name=f"{file_name_suffix}")

def main():

    full_census = choose_file("Full Census:")
    enrollment_census = choose_file("Enrollment Census:")

    # enrollment_census = "C:/Users/Brendon/Desktop/Employee Navigator/SalesVista_LLC_-_Enrollment_Census_(Vertical) (4).csv"
    # full_census = "C:/Users/Brendon/Desktop/Employee Navigator/SalesVista_LLC_-_Full_Census (1).csv"

    wb_name_enrollment_census = "temp_enrollment_census.xlsx"
    wb_name_full_census = "temp_full_census.xlsx"

    wb_enrollment_census = csv_to_excel(enrollment_census, wb_name_enrollment_census)
    wb_full_census = csv_to_excel(full_census, wb_name_full_census)

    # remove_file(f"./{wb_name_enrollment_census}")
    # remove_file(f"./{wb_name_full_census}")

    # wb_enrollment_census.save(wb_name_enrollment_census)
    # wb_full_census.save(wb_name_full_census)


    plan_types = choose_disability()

    # p1 = Process(
    main_basic_benefits_census(wb=wb_name_enrollment_census, file_name_suffix="2_Basic Enrollment Census")
        # )
    # p1.start()
    # p1.join()

    # p2 = Process(
    main_employee_dependent_census(wb=wb_name_full_census, file_name_suffix="1_Employee Dependent Census")
        # )
    # p2.start()
    # p2.join()

    main_hsa_enrollments(wb=wb_name_enrollment_census, file_name_suffix="3_HSA Enrollments")


    # p4 = Process(main_disability(wb=wb_name_enrollment_census, file_name_suffix="4_Disability Enrollments", plan_types=plan_types))
    # p4.start()
    # p4.join()

    main_life_add_critical_illness(wb=wb_name_enrollment_census, file_name_suffix="5_Life ADD and Critical Illness")


    os.remove(f"./{wb_name_enrollment_census}")
    os.remove(f"./{wb_name_full_census}")
    

if __name__ == "__main__":
    main()
